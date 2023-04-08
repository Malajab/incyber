import openpyxl
import ipaddress

# Open the input Excel file
input_file = openpyxl.load_workbook('input_file.xlsx')
input_sheet = input_file.active

# Create a new Excel file to store the results
output_file = openpyxl.Workbook()
output_sheet = output_file.active

# Write column headers to the output file
output_sheet.cell(row=1, column=1, value='CIDR')
output_sheet.cell(row=1, column=2, value='IP Count')

# Loop through the rows of the input file and search for CIDR notation IP addresses
for row in input_sheet.iter_rows(values_only=True):
    for cell in row:
        if isinstance(cell, str) and '/' in cell:
            try:
                # Calculate the number of IP addresses in the CIDR block
                ip_network = ipaddress.IPv4Network(cell, strict=False)
                ip_count = ip_network.num_addresses
                
                # Write the CIDR block and IP count to the output file
                output_sheet.append([str(ip_network), ip_count])
            except ValueError:
                # Ignore any invalid CIDR notation IP addresses
                pass

# Save the output file
output_file.save('output_file.xlsx')
